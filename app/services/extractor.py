import asyncio
import io
import logging
from typing import BinaryIO

import openpyxl  # For .xlsx files
import pdfplumber
import pytesseract

# docx for Word documents
from docx import Document
from pdf2image import convert_from_bytes
from pdf2image.exceptions import PDFInfoNotInstalledError
from pdf2image.exceptions import PDFPageCountError
from pdf2image.exceptions import PDFSyntaxError

from app.core.config import settings
from app.core.ocr import ocr as ocr_image_file_directly

# Configure module logger
logger = logging.getLogger(__name__)

# --- Configuration for PDF OCR Fallback ---
MIN_PDF_TEXT_LENGTH_FOR_DIRECT_EXTRACTION = 50  # Threshold for triggering OCR
PDF_OCR_DPI = 150  # DPI for converting PDF pages to images for OCR
# ------------------------------------------


class ExtractorError(Exception):
    """Base exception for extraction-related errors"""


async def _ocr_pdf_pages(pdf_file_bytes: bytes, fname: str, request_id: str) -> str:
    """
    Converts PDF pages to images and OCRs them.
    Runs synchronous pdf2image and PIL/pytesseract operations in a thread.
    """

    def _sync_ocr_pdf_pages_processing(file_bytes: bytes, lang: str, ocr_dpi_setting: int) -> str:
        all_pages_text_list: list[str] = []
        try:
            logger.info("[%s] PDF_OCR_HELPER: Converting PDF '%s' to images at %d DPI.", request_id, fname, ocr_dpi_setting)
            # `poppler_path=""` tells pdf2image to find Poppler in PATH
            images_from_pdf = convert_from_bytes(file_bytes, dpi=ocr_dpi_setting, poppler_path="")
        except (PDFInfoNotInstalledError, PDFPageCountError) as e:
            logger.error(
                "[%s] PDF_OCR_HELPER for '%s': Poppler utilities (pdfinfo/pdftocairo) not found or PDF issue. Ensure Poppler is installed and in PATH. Error: %s",
                request_id,
                fname,
                e,
            )
            raise ExtractorError(f"Poppler/PDF issue during PDF OCR for {fname}: {e}") from e
        except PDFSyntaxError as e:
            logger.error("[%s] PDF_OCR_HELPER for '%s': PDF syntax error. File might be corrupted. Error: %s", request_id, fname, e)
            raise ExtractorError(f"PDF syntax error during PDF OCR for {fname}: {e}") from e
        except Exception as e:  # Catch other pdf2image errors
            logger.error("[%s] PDF_OCR_HELPER for '%s': Failed to convert PDF to images. Error: %s", request_id, fname, e, exc_info=True)
            raise ExtractorError(f"PDF to image conversion failed for {fname}: {e}") from e

        if not images_from_pdf:
            logger.warning("[%s] PDF_OCR_HELPER for '%s': No images were generated from PDF.", request_id, fname)
            return ""

        logger.info("[%s] PDF_OCR_HELPER for '%s': Converted to %d image(s). Starting OCR on individual pages.", request_id, fname, len(images_from_pdf))

        for i, pil_image_page in enumerate(images_from_pdf):
            try:
                # Directly use pytesseract as this part is already inside a sync function run by to_thread
                page_text = pytesseract.image_to_string(pil_image_page, lang=lang)
                all_pages_text_list.append(page_text)
                logger.debug("[%s] PDF_OCR_HELPER for '%s': Page %d/%d OCR'd, got %d chars.", request_id, fname, i + 1, len(images_from_pdf), len(page_text))
            except Exception as e_ocr_page:
                logger.error("[%s] PDF_OCR_HELPER for '%s': Error OCR'ing page %d. Error: %s", request_id, fname, i + 1, e_ocr_page, exc_info=True)
                all_pages_text_list.append(f"\n[ERROR_OCR_PAGE_{i + 1}]\n")  # Add placeholder for problematic page

        # Join pages with a clear separator that might help LLM understand page structure
        final_ocr_text = "\n\n--- PDF Page Break (OCR) ---\n\n".join(all_pages_text_list)
        logger.info("[%s] PDF_OCR_HELPER for '%s': OCR completed. Total chars from OCR: %d", request_id, fname, len(final_ocr_text))
        return final_ocr_text

    # Run the synchronous OCR logic in a separate thread
    return await asyncio.to_thread(_sync_ocr_pdf_pages_processing, pdf_file_bytes, settings.ocr_language, PDF_OCR_DPI)


async def _pdf_to_text(f: BinaryIO, fname: str, request_id: str) -> str:
    """Extract text from PDF file. Tries direct extraction, falls back to OCR if needed."""

    def _sync_pdf_direct_extraction_with_pdfplumber(file_bytes_for_plumber: bytes) -> str:
        # This internal function runs in a thread
        buffer = io.BytesIO(file_bytes_for_plumber)
        try:
            with pdfplumber.open(buffer) as pdf:
                # Avoid calling extract_text() twice by using a more efficient approach
                page_texts = []
                for p in pdf.pages:
                    text_content = p.extract_text()
                    if text_content is not None:
                        page_texts.append(text_content)
                text_from_plumber = "\n".join(page_texts)
                logger.debug("[%s] PDF_DIRECT: pdfplumber extracted %d chars from PDF '%s'", request_id, len(text_from_plumber), fname)
                return text_from_plumber
        except Exception as e_plumber:
            logger.warning("[%s] PDF_DIRECT: pdfplumber failed for '%s': %s. Will attempt OCR fallback.", request_id, fname, str(e_plumber))
            return ""  # Return empty string to indicate failure and trigger OCR

    try:
        f.seek(0)
        # Read the file content once into memory for all subsequent operations within this function
        pdf_content_bytes = await asyncio.to_thread(f.read)

        # --- 1. Attempt Direct Text Extraction ---
        logger.info("[%s] PDF_HANDLER: Attempting direct text extraction for '%s'.", request_id, fname)
        direct_text_unstripped = await asyncio.to_thread(_sync_pdf_direct_extraction_with_pdfplumber, pdf_content_bytes)
        direct_text_stripped = direct_text_unstripped.strip()  # Use stripped version for length check logic

        # --- 2. Decide if OCR Fallback is Needed ---
        if len(direct_text_stripped) < MIN_PDF_TEXT_LENGTH_FOR_DIRECT_EXTRACTION:
            logger.warning(
                "[%s] PDF_HANDLER for '%s': Direct extraction yielded only %d usable chars (threshold: %d). Attempting OCR fallback.",
                request_id,
                fname,
                len(direct_text_stripped),
                MIN_PDF_TEXT_LENGTH_FOR_DIRECT_EXTRACTION,
            )
            try:
                # --- 3. Perform OCR Fallback ---
                ocr_fallback_text_unstripped = await _ocr_pdf_pages(pdf_content_bytes, fname, request_id)
                ocr_fallback_text_stripped = ocr_fallback_text_unstripped.strip()

                # --- 4. Choose Best Result ---
                # Prefer OCR text if it's more substantial than the (short) direct text
                if len(ocr_fallback_text_stripped) > len(direct_text_stripped):
                    logger.info(
                        "[%s] PDF_HANDLER for '%s': Using OCR fallback text (%d chars). Direct text was %d chars.",
                        request_id,
                        fname,
                        len(ocr_fallback_text_stripped),
                        len(direct_text_stripped),
                    )
                    return ocr_fallback_text_unstripped  # Return unstripped for original formatting
                else:
                    logger.info(
                        "[%s] PDF_HANDLER for '%s': OCR fallback text (%d chars) was not substantially better than direct text (%d chars). "
                        "Using direct text (or empty if both failed).",
                        request_id,
                        fname,
                        len(ocr_fallback_text_stripped),
                        len(direct_text_stripped),
                    )
                    return direct_text_unstripped  # Stick with direct_text (even if short/empty)
            except ExtractorError as e_ocr_process:  # Catch errors specifically from _ocr_pdf_pages
                logger.error("[%s] PDF_HANDLER for '%s': OCR fallback process failed: %s. Will use direct extraction result (if any).", request_id, fname, e_ocr_process)
                return direct_text_unstripped  # Fallback to whatever direct_text got if OCR itself errored

        else:
            # Direct extraction was sufficient
            logger.info("[%s] PDF_HANDLER for '%s': Direct text extraction sufficient (%d chars). OCR not attempted.", request_id, fname, len(direct_text_stripped))
            return direct_text_unstripped

    except ExtractorError:  # Re-raise ExtractorErrors if they bubbled up uncaught
        raise
    except Exception as e:  # Catch-all for other unexpected errors in this handler
        logger.error("[%s] PDF_HANDLER for '%s': An unexpected error occurred during PDF processing: %s", request_id, fname, str(e), exc_info=True)
        raise ExtractorError(f"Unexpected failure processing PDF: {fname}") from e


async def _docx_to_text(f: BinaryIO, fname: str, request_id: str) -> str:
    """Extract text from DOCX file."""

    def _sync_docx_extraction(file_bytes: bytes) -> str:
        buffer = io.BytesIO(file_bytes)
        doc = Document(buffer)
        text = "\n".join(p.text for p in doc.paragraphs)
        logger.debug("[%s] DOCX: Extracted %d chars from DOCX '%s'", request_id, len(text), fname)
        return text

    try:
        f.seek(0)
        file_bytes = await asyncio.to_thread(f.read)  # Ensure async read
        return await asyncio.to_thread(_sync_docx_extraction, file_bytes)
    except Exception as e:
        logger.error("[%s] DOCX: Failed to extract text from DOCX '%s': %s", request_id, fname, str(e), exc_info=True)
        raise ExtractorError(f"Failed to extract text from DOCX: {fname}") from e


async def _image_handler(fname: str, f: BinaryIO, request_id: str) -> str:
    """Return text from image using OCR."""
    logger.info("[%s] IMAGE_HANDLER: Processing image file: %s", request_id, fname)
    try:
        f.seek(0)
        # Use the renamed import from Step 2
        text = await ocr_image_file_directly(f)
        logger.debug("[%s] IMAGE_HANDLER: OCR for '%s' extracted %d chars", request_id, fname, len(text.strip()))

        if len(text.strip()) > 0:
            logger.info("[%s] IMAGE_HANDLER: Using OCR text for '%s' (length > 0)", request_id, fname)
            return text
        else:
            logger.info("[%s] IMAGE_HANDLER: No OCR text found for '%s'", request_id, fname)
            return ""
    except Exception as e:
        logger.exception("[%s] IMAGE_HANDLER: Failed to handle image file '%s'", request_id, fname)
        raise ExtractorError(f"Failed to handle image file: {fname}") from e


async def _excel_to_text(f: BinaryIO, filename: str, request_id: str) -> str:
    """
    Asynchronously extracts text from an Excel file by running sync extraction in a thread.
    f: A binary file-like object.
    filename: The original name of the file, used to determine .xlsx vs .xls.
    request_id: The unique request ID for logging purposes.
    """
    try:
        f.seek(0)
        file_bytes = await asyncio.to_thread(f.read)  # Read the whole file into memory for the sync function
        return await asyncio.to_thread(_sync_excel_extraction, file_bytes, filename, request_id)
    except Exception as e:
        # Log error from this async wrapper if reading bytes fails or thread call fails
        logger.error("[%s] EXCEL_ASYNC: Async wrapper for Excel extraction failed for '%s': %s", request_id, filename, str(e), exc_info=True)
        raise ExtractorError(f"Error during async processing of Excel file: {filename}") from e


def _sync_excel_extraction(file_bytes: bytes, filename: str, request_id: str) -> str:
    """
    Extracts data from all sheets of an Excel file (.xlsx or .xls)
    and represents each as CSV-formatted text, delimited by markers.
    Uses openpyxl for .xlsx and xlrd3 (if installed) for .xls.
    """
    logger.debug("[%s] EXCEL_SYNC: Attempting to extract text from Excel file: %s", request_id, filename)
    file_extension = filename.lower().split(".")[-1]
    excel_buffer = io.BytesIO(file_bytes)
    all_sheets_content: list[str] = []

    try:
        if file_extension == "xlsx":
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
            for i, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheet_lines: list[str] = [f"--- START EXCEL SHEET (File: {filename}, Sheet Index: {i}, Sheet Name: {sheet_name}) ---"]
                if sheet.max_row == 0 and sheet.max_column == 0:  # Check if sheet is truly empty
                    sheet_lines.append("(Sheet is empty)")
                else:
                    for row_tuple in sheet.iter_rows():
                        # Ensure all cell values are converted to string, handling None
                        row_values = [str(cell.value) if cell.value is not None else "" for cell in row_tuple]
                        sheet_lines.append(",".join(row_values))

                sheet_lines.append(f"--- END EXCEL SHEET (Sheet Name: {sheet_name}) ---")
                all_sheets_content.append("\n".join(sheet_lines))

        elif file_extension == "xls":
            try:
                # Try xlrd3 first, then fall back to xlrd if needed
                try:
                    from xlrd3 import XLRDError  # type: ignore
                    from xlrd3 import open_workbook  # type: ignore
                except ImportError:
                    try:
                        from xlrd import XLRDError  # type: ignore
                        from xlrd import open_workbook  # type: ignore
                    except ImportError:
                        logger.warning(
                            "[%s] EXCEL_SYNC: Neither xlrd3 nor xlrd library is installed. Cannot process .xls files like '%s'. Please install 'xlrd3' to enable .xls support.",
                            request_id,
                            filename,
                        )
                        # Return a clear message instead of raising an error immediately
                        return f"--- ERROR: Processing .xls file '{filename}' requires 'xlrd3' library, which is not installed. ---"
            except Exception as e_import:
                logger.error("[%s] EXCEL_SYNC: Error importing xlrd modules: %s", request_id, str(e_import))
                return f"--- ERROR: Processing .xls file '{filename}' encountered library import error: {e_import} ---"

            try:
                workbook = open_workbook(file_contents=file_bytes)  # xlrd uses file_contents
                for i in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(i)
                    sheet_name = sheet.name
                    xls_sheet_lines: list[str] = [f"--- START EXCEL SHEET (File: {filename}, Sheet Index: {i}, Sheet Name: {sheet_name}) ---"]
                    if sheet.nrows == 0 and sheet.ncols == 0:
                        xls_sheet_lines.append("(Sheet is empty)")
                    else:
                        for row_idx in range(sheet.nrows):
                            row_values = []
                            for col_idx in range(sheet.ncols):
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                row_values.append(str(cell_value) if cell_value is not None else "")
                            xls_sheet_lines.append(",".join(row_values))
                    xls_sheet_lines.append(f"--- END EXCEL SHEET (Sheet Name: {sheet_name}) ---")
                    all_sheets_content.append("\n".join(xls_sheet_lines))
            except XLRDError as e_xlrd:  # Catch specific xlrd errors
                logger.error("[%s] EXCEL_SYNC: xlrd failed to parse .xls file '%s': %s", request_id, filename, str(e_xlrd))
                raise ExtractorError(f"Failed to parse .xls file '{filename}' with xlrd: {e_xlrd}") from e_xlrd

        else:
            logger.warning("[%s] EXCEL_SYNC: Unsupported Excel extension for file: %s", request_id, filename)
            raise ExtractorError(f"Unsupported Excel file type for: {filename}")

        final_text = "\n\n".join(all_sheets_content)
        logger.debug("[%s] EXCEL_SYNC: Extracted %d chars (as CSVs) from Excel file: '%s'", request_id, len(final_text), filename)
        return final_text

    except Exception as e:  # Catch any other unexpected errors during processing
        logger.error("[%s] EXCEL_SYNC: Failed to extract text from Excel file '%s': %s", request_id, filename, str(e), exc_info=True)
        raise ExtractorError(f"Failed to extract text from Excel file: {filename}") from e


async def extract(fname: str, f: BinaryIO, request_id: str) -> str:
    """Extract text from a file based on its extension."""
    ext = fname.lower().split(".")[-1]
    logger.info("[%s] EXTRACT_MAIN: Starting extraction for file: '%s' (type: %s)", request_id, fname, ext)
    extracted_text = ""  # Initialize

    try:
        if ext == "pdf":
            extracted_text = await _pdf_to_text(f, fname, request_id)
        elif ext in {"docx", "doc"}:
            extracted_text = await _docx_to_text(f, fname, request_id)
        elif ext in {"xlsx", "xls"}:
            extracted_text = await _excel_to_text(f, fname, request_id)
        elif ext in {"png", "jpg", "jpeg"}:
            extracted_text = await _image_handler(fname, f, request_id)
        else:
            logger.warning(
                "[%s] EXTRACT_MAIN: Unsupported file type '%s' for file '%s'.",
                request_id,
                ext,
                fname,
            )
            raise ExtractorError(f"Unsupported file type: '{ext}' for file '{fname}'")

        # Final log after successful processing by a handler
        logger.info(
            "[%s] EXTRACT_MAIN: Successfully processed file '%s' (type: %s). Final extracted chars: %d",
            request_id,
            fname,
            ext.upper(),
            len(extracted_text.strip()),  # Use .strip() for meaningful count
        )
        return extracted_text

    except ExtractorError:  # Re-raise if a handler already raised it (it should have been logged there)
        logger.error("[%s] EXTRACT_MAIN: Extraction failed for '%s' due to a caught ExtractorError.", request_id, fname)
        raise
    except Exception as e:  # Catch any other unexpected errors at this top level
        logger.exception("[%s] EXTRACT_MAIN: Unexpected critical error during text extraction for file: '%s'", request_id, fname)
        raise ExtractorError(f"Unexpected critical failure to process file: {fname}") from e


def guard_corpus(corpus: str, request_id: str) -> str:
    """Ensure corpus doesn't exceed maximum length."""
    original_len = len(corpus)

    if original_len > settings.max_prompt_chars:
        logger.warning(
            "[%s] CORPUS_GUARD: Corpus exceeds max length (%d > %d), truncating",
            request_id,
            original_len,
            settings.max_prompt_chars,
        )
        return corpus[: settings.max_prompt_chars] + "\n\n[TESTO TRONCATO PER LIMITE TOKEN]"

    logger.debug("[%s] CORPUS_GUARD: Corpus length OK: %d chars", request_id, original_len)
    return corpus
